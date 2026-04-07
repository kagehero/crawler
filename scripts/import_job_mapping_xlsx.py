#!/usr/bin/env python3
"""
Import job_mapping_with_selection.xlsx → src/data/job-category-groups.json

Expected columns (1-based Excel): B = ジョブメドレー職種, C = ウェルミー職種, F = 画面表示ラベル（グループ名）
One row per mapping; same F value merges into one group with combined B/C lists.

openpyxl (PEP 668 / externally-managed Python):
  sudo apt install python3-openpyxl
  # or: python3 -m venv .venv && .venv/bin/pip install openpyxl && .venv/bin/python scripts/...

Usage (repo root — default output is always src/data/ under this repo):
  python3 scripts/import_job_mapping_xlsx.py path/to/job_mapping_with_selection.xlsx

From crawler/ directory:
  python3 ../scripts/import_job_mapping_xlsx.py ../path/to/job_mapping_with_selection.xlsx
  # or: python3 scripts/import_job_mapping_xlsx.py  (wrapper under crawler/scripts/)
"""
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit(
        "openpyxl is required. Try one of:\n"
        "  sudo apt install python3-openpyxl\n"
        "  python3 -m venv .venv && .venv/bin/pip install openpyxl\n"
    ) from e

_REPO_ROOT = Path(__file__).resolve().parent.parent


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="Path to job_mapping_with_selection.xlsx")
    parser.add_argument(
        "-o",
        "--output",
        default=str(_REPO_ROOT / "src/data/job-category-groups.json"),
        help="Output JSON path (default: <repo>/src/data/job-category-groups.json)",
    )
    args = parser.parse_args()

    path = Path(args.xlsx).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    if not path.is_file():
        print(f"Error: file not found: {path}", flush=True)
        return 1

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # 0-based column indices: B=1, C=2, F=5
    col_b, col_c, col_f = 1, 2, 5

    merged: dict[str, dict[str, set[str]]] = defaultdict(
        lambda: {"jobMedleyCategories": set(), "wellmeCategories": set()}
    )

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) <= col_f:
            continue
        b = row[col_b]
        c = row[col_c]
        f = row[col_f]
        label = str(f).strip() if f is not None else ""
        if not label:
            continue
        if b is not None and str(b).strip():
            merged[label]["jobMedleyCategories"].add(str(b).strip())
        if c is not None and str(c).strip():
            merged[label]["wellmeCategories"].add(str(c).strip())

    groups = []
    for label in sorted(merged.keys(), key=lambda s: s):
        g = merged[label]
        groups.append(
            {
                "label": label,
                "jobMedleyCategories": sorted(g["jobMedleyCategories"]),
                "wellmeCategories": sorted(g["wellmeCategories"]),
            }
        )

    out = Path(args.output).expanduser()
    if not out.is_absolute():
        out = (_REPO_ROOT / out).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as fp:
        json.dump({"groups": groups}, fp, ensure_ascii=False, indent=2)
        fp.write("\n")

    print(f"Wrote {len(groups)} groups to {out}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
